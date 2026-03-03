#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Baixa o CSV do PLD horário (Dados Abertos CCEE), filtra apenas o SUBMERCADO
(SUDESTE por padrão) e salva o resultado em XLSX no mesmo diretório de destino.

Baseado em: baixar_pld_ccee.py

Principais diferenças:
- Saída em .xlsx (requer openpyxl)
- Filtra somente um subsistema (default: SUDESTE)
- Pode apagar o CSV baixado após conversão (default: apaga)
- LÊ AS CONFIGURAÇÕES A PARTIR DO ARQUIVO config.ini

Uso:
  python baixar_pld_ccee_sudeste_xlsx.py

Opções:
  --submercado "SUDESTE"   (sobrescreve o config.ini)
  --keep-csv               mantém o CSV original baixado
  --no-sort                não ordena (default: ordena por MES desc, DIA desc, HORA asc)

Observação:
- Se o Python não encontrar o pacote openpyxl:
    pip install openpyxl

Autor: João Miguel
"""

import argparse
import datetime as dt
import os
import re
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
import logging
from logging import handlers, getLogger, Formatter, INFO, StreamHandler
import csv
from decimal import Decimal, InvalidOperation
import configparser

# ===================== LENDO CONFIGURAÇÕES DO .ini =====================
config = configparser.ConfigParser()
caminho_config = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Configuracoes', 'config.ini')

if not os.path.exists(caminho_config):
    print(f"[ERRO] Arquivo config.ini não encontrado em: {caminho_config}")
    sys.exit(1)

config.read(caminho_config, encoding='utf-8')

# Diretórios
DEFAULT_DEST_DIR = config.get('DIRETORIOS', 'PLD_HORARIO')

# Configurações da CCEE
DEFAULT_URL        = config.get('CCEE_DOWNLOAD', 'URL')
DEFAULT_LOG_NAME   = config.get('CCEE_DOWNLOAD', 'LOG_NAME')
MAX_RETRIES        = config.getint('CCEE_DOWNLOAD', 'MAX_RETRIES')
CONNECT_TIMEOUT    = config.getint('CCEE_DOWNLOAD', 'CONNECT_TIMEOUT')
READ_TIMEOUT       = config.getint('CCEE_DOWNLOAD', 'READ_TIMEOUT')
USER_AGENT         = config.get('CCEE_DOWNLOAD', 'USER_AGENT')
DEFAULT_SUBMERCADO = config.get('CCEE_DOWNLOAD', 'SUBMERCADO')

# Tratando Listas e Sets a partir do texto separado por vírgula
str_expected = config.get('CCEE_DOWNLOAD', 'EXPECTED_COLUMNS')
EXPECTED_COLUMNS = {x.strip() for x in str_expected.split(',')}

str_price_cols = config.get('CCEE_DOWNLOAD', 'PRICE_COL_CANDIDATES')
PRICE_COL_CANDIDATES = [x.strip() for x in str_price_cols.split(',')]

# ===================== UTILITÁRIOS =====================

def setup_logger(log_path: str) -> logging.Logger:
    logger = getLogger("pld_downloader_xlsx")
    logger.setLevel(INFO)
    logger.handlers.clear()

    sh = StreamHandler(sys.stdout)
    sh.setLevel(INFO)
    sh.setFormatter(Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(sh)

    try:
        fh = handlers.RotatingFileHandler(log_path, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
        fh.setLevel(INFO)
        fh.setFormatter(Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(fh)
    except Exception as e:
        logger.warning(f"Não foi possível criar log em '{log_path}': {e}")

    return logger


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sanitize_filename(name: str) -> str:
    name = name.strip().replace("\n", "").replace("\r", "")
    name = re.sub(r'[<>:"/\\\n?*]', "_", name)
    return name[:240]


def extract_filename_from_cd(content_disposition: str):
    if not content_disposition:
        return None

    m = re.search(r"filename\*\s*=\s*UTF-8''([^;]+)", content_disposition, re.IGNORECASE)
    if m:
        try:
            return urllib.parse.unquote(m.group(1))
        except Exception:
            return m.group(1)

    m = re.search(r'filename\s*=\s*"([^"]+)"', content_disposition, re.IGNORECASE)
    if m:
        return m.group(1)

    m = re.search(r"filename\s*=\s*([^;]+)", content_disposition, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    return None


def build_opener():
    handlers_list = [urllib.request.ProxyHandler(), urllib.request.HTTPSHandler()]
    opener = urllib.request.build_opener(*handlers_list)
    opener.addheaders = [("User-Agent", USER_AGENT)]
    return opener


def http_head(url: str, opener, timeout: int, logger: logging.Logger):
    req = urllib.request.Request(url, method="HEAD")
    try:
        with opener.open(req, timeout=timeout) as resp:
            return resp.headers
    except Exception as e:
        logger.info(f"HEAD falhou (seguiremos sem): {e}")
        return None


def determine_filename(url: str, headers, logger: logging.Logger) -> str:
    filename = None
    if headers:
        filename = extract_filename_from_cd(headers.get("Content-Disposition", ""))
        if filename:
            logger.info(f"Nome original via Content-Disposition: {filename}")

    if not filename:
        path_name = os.path.basename(urllib.parse.urlparse(url).path)
        if path_name and path_name.lower() != "content":
            filename = path_name
            logger.info(f"Nome obtido pela URL final: {filename}")

    if not filename:
        filename = f"pld_horario_{dt.datetime.now():%Y%m%d_%H%M%S}.csv"
        logger.info(f"Nome fallback aplicado: {filename}")

    return sanitize_filename(filename)


def unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    for i in range(1, 1000):
        candidate = f"{base} ({i}){ext}"
        if not os.path.exists(candidate):
            return candidate
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{ts}{ext}"


def _detect_delimiter(header_line: str) -> str:
    return ";" if header_line.count(";") >= header_line.count(",") else ","


def validate_csv_head(sample_bytes: bytes, logger: logging.Logger) -> bool:
    try:
        text = sample_bytes.decode("utf-8", errors="ignore")
        header_line = text.splitlines()[0] if text else ""
        cols = {c.strip().strip('"') for c in header_line.split(";")}
        if len(cols) < 3:
            cols = {c.strip().strip('"') for c in header_line.split(",")}
        ok = EXPECTED_COLUMNS.issubset({c.upper() for c in cols})
        if not ok:
            logger.warning(f"Validação branda: cabeçalho inesperado. Encontrado: {cols}")
        return True
    except Exception as e:
        logger.warning(f"Falha ao validar cabeçalho (seguiremos): {e}")
        return True


def stream_download(url: str, dest_path: str, opener, logger: logging.Logger) -> str:
    req = urllib.request.Request(url, method="GET")
    with opener.open(req, timeout=CONNECT_TIMEOUT) as resp:
        final_url = resp.geturl()
        headers = resp.headers
        content_length = headers.get("Content-Length")
        logger.info(f"Baixando de: {final_url} (Content-Length={content_length})")

        final_name = determine_filename(final_url, headers, logger)
        final_path = os.path.join(os.path.dirname(dest_path), final_name)
        if os.path.abspath(final_path) != os.path.abspath(dest_path):
            dest_path = final_path

        ensure_dir(os.path.dirname(dest_path))
        with tempfile.NamedTemporaryFile(prefix=".partial_", suffix=".tmp", dir=os.path.dirname(dest_path), delete=False) as tmp:
            tmp_path = tmp.name
            logger.info(f"Gravando temporário: {tmp_path}")

            chunk = resp.read(1024 * 1024)
            first_bytes = chunk
            total = 0
            while chunk:
                tmp.write(chunk)
                total += len(chunk)
                chunk = resp.read(1024 * 1024)

        logger.info(f"Download concluído ({total} bytes).")
        validate_csv_head(first_bytes or b"", logger)

        final_dest = dest_path if not os.path.exists(dest_path) else unique_path(dest_path)
        os.replace(tmp_path, final_dest)
        logger.info(f"Arquivo salvo em: {final_dest}")
        return final_dest


def download_with_retries(url: str, dest_dir: str, overwrite: bool, logger: logging.Logger):
    opener = build_opener()

    headers = http_head(url, opener, CONNECT_TIMEOUT, logger)
    filename = determine_filename(url, headers, logger)
    dest_path = os.path.join(dest_dir, filename)

    delay = 3.0
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            if os.path.exists(dest_path) and overwrite:
                logger.info(f"overwrite=True: removendo arquivo existente: {dest_path}")
                os.remove(dest_path)
            return stream_download(url, dest_path, opener, logger)
        except urllib.error.HTTPError as e:
            logger.error(f"Tentativa {attempt}/{MAX_RETRIES} - HTTPError {e.code}: {e.reason}")
        except urllib.error.URLError as e:
            logger.error(f"Tentativa {attempt}/{MAX_RETRIES} - URLError: {e.reason}")
        except Exception as e:
            logger.error(f"Tentativa {attempt}/{MAX_RETRIES} - Erro inesperado: {e!r}")

        if attempt < MAX_RETRIES:
            logger.info(f"Aguardando {delay:.1f}s para tentar novamente...")
            time.sleep(delay)
            delay *= 2

    logger.error("Falha após todas as tentativas.")
    return None


def _find_col_idx(header, name_upper: str, fallback=None):
    try:
        return header.index(name_upper)
    except ValueError:
        return fallback


def _find_price_col(header_up, preferred_name: str | None, logger: logging.Logger) -> int:
    if preferred_name:
        up = preferred_name.strip().upper()
        if up in header_up:
            idx = header_up.index(up)
            logger.info(f"Coluna de preço por parâmetro: {preferred_name} (índice {idx})")
            return idx
        logger.warning(f"Nome de coluna informado não encontrado: {preferred_name}. Tentando candidatos.")

    for cand in PRICE_COL_CANDIDATES:
        if cand in header_up:
            idx = header_up.index(cand)
            logger.info(f"Coluna de preço identificada: {cand} (índice {idx})")
            return idx

    logger.warning("Nenhum nome de coluna de preço encontrado. Usando fallback Coluna F (índice 5).")
    return 5


def _parse_decimal(value: str):
    """Converte string numérica para Decimal, detectando separadores (, .) e milhares."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None

    last_dot = s.rfind(".")
    last_comma = s.rfind(",")

    if last_dot == -1 and last_comma == -1:
        s_norm = s
    elif last_dot > last_comma:
        s_norm = s.replace(",", "")
    else:
        s_norm = s.replace(".", "").replace(",", ".")

    try:
        return Decimal(s_norm)
    except InvalidOperation:
        return None


def _sort_key(mes, dia, hora):
    # MES desc, DIA desc, HORA asc
    return (-mes, -dia, hora)


def convert_csv_to_xlsx_filtered(
    csv_path: str,
    xlsx_path: str,
    logger: logging.Logger,
    submercado: str = "SUDESTE",
    price_col_name: str | None = "PLD_HORA",
    do_sort: bool = True,
):
    """Lê CSV, filtra SUBMERCADO, (opcional) ordena, e grava XLSX."""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as e:
        raise RuntimeError(
            "Pacote 'openpyxl' não encontrado. Instale com: pip install openpyxl"
        ) from e

    logger.info(f"Convertendo CSV -> XLSX (filtrando SUBMERCADO={submercado}): {csv_path} -> {xlsx_path}")

    # Detecta delimitador pela 1ª linha
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        first_chunk = f.read(4096)
        if not first_chunk:
            raise RuntimeError("CSV vazio.")
        first_line = first_chunk.splitlines()[0]
        delimiter = _detect_delimiter(first_line)

    rows = []
    with open(csv_path, "r", encoding="utf-8", newline="") as fin:
        rdr = csv.reader(fin, delimiter=delimiter)
        try:
            header = next(rdr)
        except StopIteration:
            raise RuntimeError("CSV sem linhas.")

        header_up = [h.strip().strip('"').upper() for h in header]

        idx_mes = _find_col_idx(header_up, "MES_REFERENCIA")
        idx_dia = _find_col_idx(header_up, "DIA")
        idx_hora = _find_col_idx(header_up, "HORA")
        idx_sub = _find_col_idx(header_up, "SUBMERCADO")
        if None in (idx_mes, idx_dia, idx_hora, idx_sub):
            raise RuntimeError("Cabeçalho precisa conter MES_REFERENCIA, DIA, HORA e SUBMERCADO.")

        price_idx = _find_price_col(header_up, price_col_name, logger)

        wanted = submercado.strip().upper()

        def to_int(v, default=0):
            try:
                return int(str(v).strip())
            except Exception:
                return default

        for row in rdr:
            if idx_sub >= len(row):
                continue
            sub = str(row[idx_sub]).strip().strip('"').upper()
            if sub != wanted:
                continue

            mes = to_int(row[idx_mes]) if idx_mes < len(row) else 0
            dia = to_int(row[idx_dia]) if idx_dia < len(row) else 0
            hora = to_int(row[idx_hora]) if idx_hora < len(row) else 0

            # Converte preço para Decimal/float (se existir)
            if price_idx < len(row):
                d = _parse_decimal(row[price_idx])
                if d is not None:
                    # Mantém 2 casas
                    d = d.quantize(Decimal("0.01"))
                    row[price_idx] = float(d)

            rows.append((mes, dia, hora, row))

    if do_sort:
        rows.sort(key=lambda t: _sort_key(t[0], t[1], t[2]))

    # Monta workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"PLD_{wanted}"

    # Cabeçalho
    ws.append(header)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # Dados
    for _, __, ___, row in rows:
        ws.append(row)

    # Formatação simples
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ajuste simples de largura (limite)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[: min(len(col), 5000)]:
            val = cell.value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    # Número com 2 casas para coluna de preço (se for achada)
    if 0 <= price_idx < len(header):
        # openpyxl usa 1-based para coluna
        col_num = price_idx + 1
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=col_num)
            if isinstance(c.value, (int, float)):
                c.number_format = "0.00"

    # Escrita atômica
    out_dir = os.path.dirname(xlsx_path) or "."
    ensure_dir(out_dir)
    with tempfile.NamedTemporaryFile(prefix=".xlsx_", suffix=".tmp", dir=out_dir, delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        # Evita sobrescrever
        final_dest = xlsx_path if not os.path.exists(xlsx_path) else unique_path(xlsx_path)
        os.replace(tmp_path, final_dest)
        logger.info(f"XLSX salvo em: {final_dest}")
        return final_dest
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def rename_to_standard_xlsx(xlsx_path: str, logger: logging.Logger, prefix: str = "preco_horario_sudeste - ") -> str:
    now = dt.datetime.now()
    ts = now.strftime("%Y-%m-%dT%H%M%S") + f".{int(now.microsecond/1000):03d}"
    new_name = sanitize_filename(f"{prefix}{ts}.xlsx")
    new_path = os.path.join(os.path.dirname(xlsx_path), new_name)
    if os.path.exists(new_path):
        new_path = unique_path(new_path)
    os.replace(xlsx_path, new_path)
    logger.info(f"Arquivo renomeado para: {new_path}")
    return new_path


# ===================== MAIN =====================

def main():
    parser = argparse.ArgumentParser(description="Baixador do PLD horário (Dados Abertos CCEE) -> XLSX filtrado por subsistema")
    parser.add_argument("--url", default=DEFAULT_URL, help="URL do recurso (/content) no Dados Abertos")
    parser.add_argument("--dest", default=DEFAULT_DEST_DIR, help="Pasta de destino")
    parser.add_argument("--overwrite", action="store_true", help="Permite sobrescrever arquivo existente")
    parser.add_argument("--log", default=None, help="Caminho do arquivo de log (padrão: dentro da pasta de destino)")

    # Alterado de "SUDESTE" fixo para puxar a variável carregada do config.ini
    parser.add_argument("--submercado", default=DEFAULT_SUBMERCADO, help=f"SUBMERCADO a manter (padrão: {DEFAULT_SUBMERCADO})")
    
    parser.add_argument("--price-col-name", default="PLD_HORA", help="Nome da coluna de preço (padrão: PLD_HORA)")
    parser.add_argument("--keep-csv", action="store_true", help="Mantém o CSV baixado (por padrão apaga após conversão)")
    parser.add_argument("--no-sort", action="store_true", help="Não ordenar (padrão: ordena MES desc, DIA desc, HORA asc)")
    parser.add_argument("--no-rename", action="store_true", help="Não renomear com o padrão 'preco_horario_sudeste - <timestamp>.xlsx'")

    args = parser.parse_args()

    dest_dir = args.dest
    ensure_dir(dest_dir)

    log_path = args.log or os.path.join(dest_dir, DEFAULT_LOG_NAME)
    logger = setup_logger(log_path)

    logger.info("=" * 80)
    logger.info("Iniciando download do PLD horário (Dados Abertos CCEE) -> XLSX filtrado (Config.ini Integrado)")
    logger.info(f"URL: {args.url}")
    logger.info(f"Destino: {dest_dir}")
    logger.info(f"Submercado: {args.submercado}")
    logger.info(f"Overwrite: {args.overwrite}")
    logger.info(f"Ordenar: {not args.no_sort}")
    logger.info(f"Manter CSV: {args.keep_csv}")

    if re.match(r"^[A-Za-z]:", dest_dir) and not os.path.isdir(dest_dir):
        logger.warning(
            "A pasta de destino não existe (letra de unidade?). "
            "Se for unidade de rede mapeada, considere usar caminho UNC (\\\\servidor\\pasta\\...)."
        )

    csv_path = download_with_retries(args.url, dest_dir, args.overwrite, logger)
    if not csv_path:
        logger.error("Download não realizado. Encerrando com erro.")
        sys.exit(1)

    # Define nome de saída (.xlsx) no mesmo diretório
    base_name = os.path.splitext(os.path.basename(csv_path))[0]
    wanted = args.submercado.strip().upper()
    out_name = sanitize_filename(f"{base_name}_{wanted}.xlsx")
    xlsx_path = os.path.join(dest_dir, out_name)

    try:
        final_xlsx = convert_csv_to_xlsx_filtered(
            csv_path=csv_path,
            xlsx_path=xlsx_path,
            logger=logger,
            submercado=wanted,
            price_col_name=args.price_col_name if args.price_col_name.strip() else None,
            do_sort=(not args.no_sort),
        )

        if not args.keep_csv:
            try:
                os.remove(csv_path)
                logger.info(f"CSV removido (keep-csv=False): {csv_path}")
            except Exception as e:
                logger.warning(f"Não foi possível remover o CSV: {e}")

        if not args.no_rename:
            # Se o usuário trocar submercado, ainda usa prefixo com nome do submercado.
            prefix = f"preco_horario_{wanted.lower()} - "
            final_xlsx = rename_to_standard_xlsx(final_xlsx, logger, prefix=prefix)

        logger.info(f"Concluído com sucesso. Arquivo final: {final_xlsx}")
        sys.exit(0)

    except Exception as e:
        logger.error(f"Falha na conversão para XLSX: {e!r}")
        sys.exit(1)


if __name__ == "__main__":
    main()