#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera a planilha diária de Modulação (Consumo e Cessão) a partir de um template,
preenchendo entradas e PLD horário do SUDESTE.

Novidades:
- Nome do arquivo gerado inclui HORA (HHMMSS) ao lado da data para evitar duplicidade.
- Log (txt) gravado por padrão na pasta de logs definida no config.ini.
- Executa um script VBS ao final (recalcular/salvar/fechar), com caminho padrão configurado.
- Suporte a Template alternativo para Fins de Semana controlado via config.ini.

Regra de data-alvo (janela do PLD):
- A CCEE só disponibiliza o PLD do dia seguinte após ~17h.
    * Se horário local >= 17:00 -> data-alvo = amanhã
    * Se horário local <  17:00 -> data-alvo = hoje

Autor: João Miguel
"""

import argparse
import datetime as dt
import os
import re
import sys
import tempfile
import logging
import subprocess
import shlex
import configparser
from logging import handlers, getLogger, Formatter, INFO, StreamHandler


# ===================== LENDO CONFIGURAÇÕES DO .ini =====================
caminho_config = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Configuracoes', 'config.ini')

if not os.path.exists(caminho_config):
    # Fallback caso alguém rode fora do orquestrador
    caminho_config = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.ini')
    if not os.path.exists(caminho_config):
        print(f"[ERRO] Arquivo config.ini não encontrado!")
        sys.exit(1)

config = configparser.ConfigParser()
config.read(caminho_config, encoding='utf-8')

TEMPLATE_PATH    = config.get('DIRETORIOS', 'TEMPLATE_PLANILHA')
DEST_DIR_DEFAULT = config.get('DIRETORIOS', 'SAIDA_PLANILHAS')
PLD_DIR_DEFAULT  = config.get('DIRETORIOS', 'PLD_HORARIO')
LOGS_DIR_DEFAULT = config.get('DIRETORIOS', 'DIRETORIO_LOGS')
VBS_PATH_DEFAULT = config.get('DIRETORIOS', 'VBS_SCRIPT')
LOG_NAME_DEFAULT = "gerar_modulacao_parada_diaria.txt"

# Valores mensais
str_consumo = config.get('REGRAS_NEGOCIO', 'CONSUMO_MEDIO_MWM_MES')
CONSUMO_MEDIO_MWM = [float(x.strip()) for x in str_consumo.split(',')]

str_recurso = config.get('REGRAS_NEGOCIO', 'TOTAL_RECURSO_MES')
TOTAL_RECURSO = [float(x.strip()) for x in str_recurso.split(',')]

SUBMERCADO_WANTED = config.get('CCEE_DOWNLOAD', 'SUBMERCADO')

# Flag de Fim de Semana (retorna False se a chave não existir)
USAR_TEMPLATE_FIM_DE_SEMANA = config.getboolean('REGRAS_NEGOCIO', 'USAR_TEMPLATE_FIM_DE_SEMANA', fallback=False)

# ===================== LOG =====================
def setup_logger(log_path: str) -> logging.Logger:
    logger = getLogger("modulacao_diaria")
    logger.setLevel(INFO)
    logger.handlers.clear()

    sh = StreamHandler(sys.stdout)
    sh.setLevel(INFO)
    sh.setFormatter(Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(sh)

    try:
        fh = handlers.RotatingFileHandler(
            log_path,
            maxBytes=5 * 1024 * 1024,
            backupCount=3,
            encoding="utf-8",
        )
        fh.setLevel(INFO)
        fh.setFormatter(Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(fh)
    except Exception as e:
        logger.warning(f"Não foi possível criar log em '{log_path}': {e}")

    return logger


# ===================== UTIL =====================
def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def sanitize_filename(name: str) -> str:
    name = name.strip().replace("\n", "").replace("\r", "")
    name = re.sub(r'[<>:"/\\\n?*]', "_", name)
    return name[:240]


def atomic_copy(src: str, dst: str):
    """Copia src para dst de forma segura (escreve em tmp e renomeia)."""
    dst_dir = os.path.dirname(dst)
    ensure_dir(dst_dir)

    with open(src, "rb") as fsrc:
        with tempfile.NamedTemporaryFile(
            prefix=".partial_",
            suffix=".tmp",
            dir=dst_dir,
            delete=False,
        ) as tmp:
            tmp_path = tmp.name
            while True:
                chunk = fsrc.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)

    os.replace(tmp_path, dst)


def run_vbs(vbs_path: str, logger: logging.Logger, vbs_args=None, engine: str = "cscript", timeout: int = 300, cwd: str = None) -> int:
    """Executa um script .vbs via cscript/wscript (Windows)."""
    if os.name != "nt":
        logger.warning("Execução de VBS ignorada: este ambiente não parece ser Windows (os.name != 'nt').")
        return 0

    if not vbs_path:
        logger.info("VBS não informado. Pulando etapa VBS.")
        return 0

    if not os.path.isfile(vbs_path):
        logger.error(f"Script VBS não encontrado: {vbs_path}")
        return 1

    engine = (engine or "cscript").lower().strip()
    if engine not in ("cscript", "wscript"):
        logger.warning(f"Engine VBS desconhecido '{engine}', usando 'cscript'.")
        engine = "cscript"

    exe = "cscript.exe" if engine == "cscript" else "wscript.exe"
    cmd = [exe, "//nologo", vbs_path]

    if vbs_args:
        cmd.extend([str(a) for a in vbs_args])

    logger.info("Executando VBS: " + " ".join(cmd))

    try:
        completed = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, cwd=cwd)
    except FileNotFoundError as e:
        logger.error(f"Não foi possível executar '{exe}'. Erro: {e}")
        return 1
    except subprocess.TimeoutExpired:
        logger.error(f"VBS excedeu o timeout de {timeout}s e foi interrompido.")
        return 1
    except Exception as e:
        logger.error(f"Falha ao executar VBS: {e!r}")
        return 1

    if completed.stdout and completed.stdout.strip():
        logger.info("VBS stdout:\n" + completed.stdout.strip())
    if completed.stderr and completed.stderr.strip():
        logger.warning("VBS stderr:\n" + completed.stderr.strip())
    if completed.returncode != 0:
        logger.error(f"VBS retornou código {completed.returncode}.")

    return int(completed.returncode)


def list_pld_xlsx_candidates(pld_dir: str):
    """Lista candidatos de PLD XLSX, ordenados por mtime desc."""
    if not os.path.isdir(pld_dir):
        return []

    scored = []
    for fn in os.listdir(pld_dir):
        if not fn.lower().endswith(".xlsx"):
            continue

        full = os.path.join(pld_dir, fn)
        try:
            mtime = os.path.getmtime(full)
        except Exception:
            continue

        name = fn.lower()
        bonus = 1 if ("preco_horario_sudeste" in name or "pld_sudeste" in name) else 0
        scored.append((bonus, mtime, full))

    scored.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return [t[2] for t in scored]


def load_pld_for_date(pld_xlsx: str, target_date: dt.date, logger: logging.Logger):
    """Lê o XLSX do PLD e retorna lista com 24 valores para a data-alvo."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("Pacote 'openpyxl' não encontrado. Instale com: pip install openpyxl") from e

    logger.info(f"Lendo PLD do arquivo: {pld_xlsx}")

    wb = load_workbook(pld_xlsx, data_only=False, read_only=True)
    ws = wb.worksheets[0]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [(str(cell).strip().strip('"').upper()) if cell is not None else "" for cell in header_row]

    def idx(col_name: str):
        try: return header.index(col_name)
        except ValueError: return None

    idx_mes = idx("MES_REFERENCIA")
    idx_dia = idx("DIA")
    idx_hora = idx("HORA")
    idx_pld = idx("PLD_HORA")
    idx_sub = idx("SUBMERCADO")

    if None in (idx_mes, idx_dia, idx_hora, idx_pld):
        wb.close()
        raise RuntimeError("Não encontrei colunas necessárias no PLD XLSX. (MES_REFERENCIA, DIA, HORA, PLD_HORA)")

    target_mes_ref = target_date.year * 100 + target_date.month
    wanted_sub = SUBMERCADO_WANTED

    by_hour = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        try: mes_ref = int(str(row[idx_mes]).strip()) if row[idx_mes] is not None else None
        except Exception: mes_ref = None

        if mes_ref != target_mes_ref: continue

        try: dia = int(str(row[idx_dia]).strip()) if row[idx_dia] is not None else None
        except Exception: continue

        if dia != target_date.day: continue

        try: hora = int(str(row[idx_hora]).strip()) if row[idx_hora] is not None else None
        except Exception: continue

        if hora is None or hora < 0 or hora > 23: continue

        if idx_sub is not None:
            sub = str(row[idx_sub]).strip().strip('"').upper() if row[idx_sub] is not None else ""
            if sub and sub != wanted_sub: continue

        val = row[idx_pld]
        if val is None: continue

        try: fval = float(val)
        except Exception:
            try: fval = float(str(val).replace(".", "").replace(",", "."))
            except Exception: continue

        by_hour[hora] = fval

    wb.close()

    missing = [h for h in range(24) if h not in by_hour]
    if missing:
        raise RuntimeError(f"PLD incompleto para {target_date:%Y-%m-%d} no arquivo '{os.path.basename(pld_xlsx)}'. Horas faltando: {missing}")

    return [by_hour[h] for h in range(24)]


def fill_template(out_xlsx: str, target_date: dt.date, pld_vals, logger: logging.Logger):
    """Abre a planilha de saída e preenche células conforme regra."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("Pacote 'openpyxl' não encontrado. Instale com: pip install openpyxl") from e

    wb = load_workbook(out_xlsx)
    ws = wb.worksheets[0]

    month = target_date.month
    consumo = CONSUMO_MEDIO_MWM[month - 1]
    recurso = TOTAL_RECURSO[month - 1]

    logger.info(f"Preenchendo mês={month:02d}: B1(consumo)={consumo}  B2(recurso)={recurso}")

    ws["B1"].value = consumo
    ws["B2"].value = recurso

    day = target_date.day
    for r in range(6, 30):
        ws[f"A{r}"].value = day

    for i in range(24):
        ws[f"C{6 + i}"].value = pld_vals[i]

    out_dir = os.path.dirname(out_xlsx) or "."
    with tempfile.NamedTemporaryFile(prefix=".save_", suffix=".tmp", dir=out_dir, delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)
        wb.close()
        os.replace(tmp_path, out_xlsx)
        logger.info(f"Planilha preenchida e salva: {out_xlsx}")
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def resolve_target_date(now: dt.datetime, cutoff_hour: int) -> dt.date:
    cutoff = dt.time(hour=cutoff_hour, minute=0)
    if now.time() >= cutoff:
        return now.date() + dt.timedelta(days=1)
    return now.date()


def main():
    parser = argparse.ArgumentParser(description="Gera planilha diária a partir do template e PLD SUDESTE (janela 17h)")

    parser.add_argument("--template", default=TEMPLATE_PATH, help="Caminho do template padrão")
    parser.add_argument("--dest-dir", default=DEST_DIR_DEFAULT, help="Pasta destino")
    parser.add_argument("--pld-xlsx", default=None, help="Caminho explícito do XLSX de PLD SUDESTE")
    parser.add_argument("--pld-dir", default=PLD_DIR_DEFAULT, help="Pasta onde procurar o XLSX de PLD")

    # Logs
    parser.add_argument("--logs-dir", default=LOGS_DIR_DEFAULT, help="Pasta onde salvar o TXT de logs")
    parser.add_argument("--log", default=None, help="Caminho do log (sobrescreve --logs-dir)")

    # Data
    parser.add_argument("--date", default=None, help="Data-alvo no formato YYYY-MM-DD")
    parser.add_argument("--cutoff-hour", type=int, default=17, help="Hora de corte (0-23). Default: 17")

    # VBS
    parser.add_argument("--vbs", dest="vbs_path", default=VBS_PATH_DEFAULT)
    parser.add_argument("--skip-vbs", action="store_true")
    parser.add_argument("--vbs-engine", default="cscript", choices=["cscript", "wscript"])
    parser.add_argument("--vbs-timeout", type=int, default=300)
    parser.add_argument("--vbs-args", default=None)
    parser.add_argument("--vbs-pass-date", action="store_true")

    args = parser.parse_args()

    # Logger
    ensure_dir(args.logs_dir)
    log_path = args.log or os.path.join(args.logs_dir, LOG_NAME_DEFAULT)
    logger = setup_logger(log_path)

    now = dt.datetime.now()

    # Define data-alvo
    if args.date:
        target_date = dt.datetime.strptime(args.date, "%Y-%m-%d").date()
        logger.info(f"Data-alvo definida por parâmetro --date: {target_date:%Y-%m-%d}")
    else:
        target_date = resolve_target_date(now, args.cutoff_hour)

    # ===================== LÓGICA DE FIM DE SEMANA =====================
    # 5 = Sábado, 6 = Domingo
    if USAR_TEMPLATE_FIM_DE_SEMANA and target_date.weekday() >= 5:
        logger.info("Fim de semana detectado e a chave USAR_TEMPLATE_FIM_DE_SEMANA está ativada!")
        
        pasta_templates = os.path.dirname(args.template)
        template_fds = os.path.join(pasta_templates, 'AAAA.MM.DD_Modulacao_Consumo e Cessao - FimDeSemana.xlsx')
        
        if os.path.isfile(template_fds):
            args.template = template_fds
            logger.info(f"Template redirecionado para: {os.path.basename(args.template)}")
        else:
            logger.warning(f"Template de fim de semana não encontrado! Usando o padrão. Verifique a pasta: {pasta_templates}")
    # ===================================================================

    ensure_dir(args.dest_dir)

    logger.info("=" * 80)
    logger.info("Gerando planilha diária de Modulação")
    logger.info(f"Agora (local): {now:%Y-%m-%d %H:%M:%S}")
    logger.info(f"Cutoff-hour: {args.cutoff_hour:02d}:00")
    logger.info(f"Data-alvo: {target_date:%Y-%m-%d}")
    logger.info(f"Template Escolhido: {args.template}")
    logger.info(f"Destino: {args.dest_dir}")
    logger.info(f"Log: {log_path}")

    if not os.path.isfile(args.template):
        logger.error(f"Template não encontrado: {args.template}")
        sys.exit(1)

    # Nome do arquivo de saída inclui hora (HHMMSS)
    time_tag = now.strftime("%H%M%S")
    out_name = sanitize_filename(f"{target_date:%Y.%m.%d}_{time_tag}_Modulacao_Consumo e Cessao.xlsx")
    out_path = os.path.join(args.dest_dir, out_name)

    if os.path.exists(out_path):
        base, ext = os.path.splitext(out_path)
        i = 1
        while os.path.exists(f"{base} ({i}){ext}"):
            i += 1
        out_path = f"{base} ({i}){ext}"

    logger.info(f"Criando cópia do template: {out_path}")
    atomic_copy(args.template, out_path)

    # Seleciona PLD XLSX
    pld_vals = None
    chosen_pld_file = None

    if args.pld_xlsx:
        if not os.path.isfile(args.pld_xlsx):
            logger.error(f"PLD XLSX não encontrado: {args.pld_xlsx}")
            sys.exit(1)
        try:
            pld_vals = load_pld_for_date(args.pld_xlsx, target_date, logger)
            chosen_pld_file = args.pld_xlsx
        except Exception as e:
            logger.error(f"Falha ao ler PLD do arquivo informado: {e}")
            sys.exit(1)
    else:
        candidates = list_pld_xlsx_candidates(args.pld_dir)
        if not candidates:
            logger.error("Não encontrei nenhum XLSX de PLD. Aguarde o download da CCEE.")
            sys.exit(1)

        last_error = None
        for cand in candidates:
            try:
                pld_vals = load_pld_for_date(cand, target_date, logger)
                chosen_pld_file = cand
                break
            except Exception as e:
                last_error = e
                logger.info(f"Candidato não serviu ({os.path.basename(cand)}): {e}")

        if pld_vals is None:
            logger.error(f"Nenhum arquivo de PLD contém o dia {target_date:%Y-%m-%d}. Último erro: {last_error}")
            sys.exit(1)

    logger.info(f"PLD escolhido: {chosen_pld_file}")

    try:
        fill_template(out_path, target_date, pld_vals, logger)
    except Exception as e:
        logger.error(f"Falha ao preencher/salvar planilha: {e!r}")
        sys.exit(1)

    # ===================== CHAMA VBS NO FINAL =====================
    if not args.skip_vbs:
        vbs_extra_args = [out_path]  
        if args.vbs_pass_date:
            vbs_extra_args.append(target_date.strftime("%Y-%m-%d"))
        if args.vbs_args:
            vbs_extra_args.extend(shlex.split(args.vbs_args, posix=False))

        rc = run_vbs(
            args.vbs_path,
            logger,
            vbs_args=vbs_extra_args,
            engine=args.vbs_engine,
            timeout=args.vbs_timeout,
        )
        if rc != 0:
            logger.error("Processo finalizado com erro na etapa VBS.")
            sys.exit(rc)

        logger.info("VBS executado com sucesso.")

    logger.info(f"OK. Arquivo final: {out_path}")
    sys.exit(0)


if __name__ == "__main__":
    main()