import os
import sys
import pandas as pd
import requests
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import configparser
import subprocess

# =============================
# CONFIGURAÇÕES
# =============================
config = configparser.ConfigParser()
caminho_config = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Configuracoes', 'config.ini')

if not os.path.exists(caminho_config):
    sys.exit(f"[ERRO] Arquivo config.ini não encontrado em: {caminho_config}")

config.read(caminho_config, encoding='utf-8')

diretorio = config.get('DIRETORIOS', 'SAIDA_PLANILHAS')
TEMP_IMG  = config.get('DIRETORIOS', 'TEMP_IMG')
VBS_SCRIPT = config.get('DIRETORIOS', 'VBS_SCRIPT') # <--- PUXANDO O VBS DO INI

TOKEN_TELEGRAM = config.get('TELEGRAM', 'BOT_TOKEN')
CHAT_ID        = config.get('TELEGRAM', 'CHAT_ID')

nome_aba             = config.get('REGRAS_NEGOCIO', 'NOME_ABA')
CONSUMO_REDUZIDO_MWM = config.getfloat('REGRAS_NEGOCIO', 'CONSUMO_REDUZIDO_MWM')

# Paleta de Cores
COR_TEXTO_ESCURO = '#084851'
COR_BOTAO_ACAO = '#008080'
COR_DESTAQUE = '#119CB0'
COR_FUNDO_APP = '#ffffff'
COR_ALERTA = '#D32F2F'
COR_MEDIA = '#FF8C00'

# Intervalo alvo
lin_inicio = 5  # linha 6
lin_fim = 29    # linha 29

# Índices das colunas (A=0, B=1, C=2...)
col_B, col_C, col_D, col_E, col_F = 1, 2, 3, 4, 5
col_I = 8  # Geração Modulada (Coluna I)
col_L, col_M, col_N = 11, 12, 13

# =============================
# UTILITÁRIOS
# =============================
def achar_arquivo_mais_recente(pasta: str) -> str:
    if not os.path.isdir(pasta):
        sys.exit(f"[ERRO] Diretório não existe: {pasta}")

    candidatos = [
        os.path.join(pasta, f) for f in os.listdir(pasta)
        if f.lower().endswith((".xlsx", ".xlsm", ".xls")) and not f.startswith("~$")
    ]
    if not candidatos:
        sys.exit("[ERRO] Nenhum arquivo Excel encontrado.")

    import re

    padrao = re.compile(
        r"^(?P<y>\d{4})\.(?P<m>\d{2})\.(?P<d>\d{2})_(?P<h>\d{2})(?P<mi>\d{2})(?P<s>\d{2})_Modulacao_Consumo",
        re.IGNORECASE
    )

    candidatos_padrao = []
    for caminho in candidatos:
        nome = os.path.basename(caminho)
        m = padrao.search(nome)
        if not m:
            continue
        try:
            dt_nome = datetime(
                int(m.group("y")), int(m.group("m")), int(m.group("d")),
                int(m.group("h")), int(m.group("mi")), int(m.group("s"))
            )
            candidatos_padrao.append((dt_nome, os.path.getmtime(caminho), caminho))
        except Exception:
            pass

    if candidatos_padrao:
        return max(candidatos_padrao, key=lambda t: (t[0], t[1]))[2]

    return max(candidatos, key=os.path.getmtime)


def to_num(series):
    return pd.to_numeric(series, errors="coerce")


# =============================
# FUNÇÕES: GRÁFICO E TELEGRAM
# =============================
def gerar_grafico_pld(horas, pld, media_pld, caminho_saida, destaques_sorted, data_relatorio):
    print("📊 Gerando gráfico do PLD Horário...")
    fig, ax = plt.subplots(figsize=(10, 6), facecolor=COR_FUNDO_APP)
    ax.set_facecolor(COR_FUNDO_APP)

    lista_horas = pd.to_numeric(horas, errors='coerce').fillna(0).astype(int).tolist()
    lista_pld = pd.to_numeric(pld, errors='coerce').fillna(0).astype(float).tolist()

    ax.plot(
        lista_horas, lista_pld,
        color=COR_DESTAQUE, marker='o', linestyle='-',
        linewidth=2.5, markersize=6, label='PLD Horário'
    )

    (h1, p1), (h2, p2) = destaques_sorted

    if p1 <= p2:
        h_min, p_min = h1, p1
        h_max, p_max = h2, p2
    else:
        h_min, p_min = h2, p2
        h_max, p_max = h1, p1

    bbox_padrao = dict(boxstyle="round,pad=0.25", facecolor=COR_FUNDO_APP, edgecolor="none", alpha=0.90)
    seta = dict(arrowstyle="-", color="gray", lw=1.0, alpha=0.6)

    ax.scatter([h_min], [p_min], color='#2E7D32', s=150, zorder=6, edgecolors=COR_FUNDO_APP, linewidths=2, label='Mín do Dia')
    ax.annotate(f'R$ {p_min:.2f}', xy=(h_min, p_min), xytext=(0, 18), textcoords="offset points", ha='center', va='bottom', fontsize=10, color='#2E7D32', weight='bold', bbox=bbox_padrao, arrowprops=seta, zorder=7)

    ax.scatter([h_max], [p_max], color=COR_ALERTA, s=150, zorder=6, edgecolors=COR_FUNDO_APP, linewidths=2, label='Máx do Dia')
    ax.annotate(f'R$ {p_max:.2f}', xy=(h_max, p_max), xytext=(0, -22), textcoords="offset points", ha='center', va='top', fontsize=10, color=COR_ALERTA, weight='bold', bbox=bbox_padrao, arrowprops=seta, zorder=7)

    ax.axhline(y=media_pld, color=COR_MEDIA, linestyle='--', linewidth=2.5, alpha=1.0, label=f'Média (R$ {media_pld:.2f})')

    ax.set_title('Variação do PLD Horário', fontsize=16, weight='bold', color=COR_TEXTO_ESCURO, pad=20)
    ax.set_xlabel('Hora do Dia', fontsize=12, color=COR_TEXTO_ESCURO, weight='bold')
    ax.set_ylabel('Valor do PLD (R$/MWh)', fontsize=12, color=COR_TEXTO_ESCURO, weight='bold')

    ax.set_xticks(lista_horas)
    ax.tick_params(colors=COR_TEXTO_ESCURO)

    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color(COR_BOTAO_ACAO)
    ax.spines['bottom'].set_color(COR_BOTAO_ACAO)

    ax.grid(True, axis='y', linestyle='--', alpha=0.3, color=COR_BOTAO_ACAO)
    ax.grid(False, axis='x')

    ax.legend(loc='upper left', frameon=True, facecolor=COR_FUNDO_APP, edgecolor=COR_BOTAO_ACAO, labelcolor=COR_TEXTO_ESCURO)
    ax.text(1.0, 1.02, f"Referência: {data_relatorio}", transform=ax.transAxes, fontsize=9, color='gray', style='italic', ha='right', va='bottom')

    plt.tight_layout()
    plt.savefig(caminho_saida, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close()
    print(f"✅ Gráfico salvo em: {caminho_saida}")


def enviar_telegram_com_foto(texto_html: str, caminho_imagem: str, token: str, chat_id: str) -> bool:
    if not token or not chat_id: return False
    url = f"https://api.telegram.org/bot{token}/sendPhoto"
    payload = {"chat_id": chat_id, "caption": texto_html, "parse_mode": "HTML"}
    try:
        with open(caminho_imagem, 'rb') as img_file:
            print("📤 Enviando mensagem e gráfico para o Telegram...")
            resp = requests.post(url, data=payload, files={'photo': img_file}, timeout=30)
            resp.raise_for_status()
            return True
    except requests.RequestException as e:
        print(f"[ERRO Telegram] {e}")
        return False

def enviar_telegram_documento(caminho_arquivo: str, token: str, chat_id: str) -> bool:
    if not token or not chat_id: return False
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    payload = {"chat_id": chat_id}
    try:
        with open(caminho_arquivo, 'rb') as doc_file:
            print("📤 Enviando planilha para o Telegram...")
            resp = requests.post(url, data=payload, files={'document': doc_file}, timeout=60)
            resp.raise_for_status()
            return True
    except requests.RequestException as e:
        print(f"[ERRO Telegram Documento] {e}")
        return False


# =============================
# LÓGICA PRINCIPAL
# =============================
def main():
    print("--- Iniciando Processamento ---\n")

    caminho = achar_arquivo_mais_recente(diretorio)
    nome_arquivo = os.path.basename(caminho)
    print(f"Lendo: {nome_arquivo}...")

    try:
        df = pd.read_excel(caminho, sheet_name=nome_aba, header=None)
    except Exception as e:
        sys.exit(f"[ERRO] Falha ao ler Excel: {e}")

    # =============================
    # MÉTRICAS BASE (FLAT)
    # =============================
    pld = to_num(df.iloc[lin_inicio:lin_fim, col_C]).fillna(0)
    horas = df.iloc[lin_inicio:lin_fim, col_B]

    res_fin = to_num(df.iloc[lin_inicio:lin_fim, col_L]).sum(min_count=1) or 0.0
    soma_D = to_num(df.iloc[lin_inicio:lin_fim, col_D]).sum(min_count=1) or 0.0
    soma_F = to_num(df.iloc[lin_inicio:lin_fim, col_F]).sum(min_count=1) or 0.0

    denominador_flat = soma_D + soma_F
    custo_mwm_flat = (res_fin / denominador_flat) if denominador_flat else 0.0

    idx_max = pld.idxmax()
    idx_min = pld.idxmin()
    media_pld = pld.mean()

    val_max = float(pld.loc[idx_max])
    val_min = float(pld.loc[idx_min])

    hora_max = int(horas.loc[idx_max])
    hora_min = int(horas.loc[idx_min])

    # =============================
    # CÁLCULO DOS CENÁRIOS (3h e 4h)
    # =============================
    consumo_medio_dia = float(to_num(df.iloc[0, col_B]))
    geracao_mod = to_num(df.iloc[lin_inicio:lin_fim, col_I]).fillna(0)
    cessao = to_num(df.iloc[lin_inicio:lin_fim, col_F]).fillna(0)

    top4_idx = pld.nlargest(4).index
    top3_idx = top4_idx[:3]

    def montar_texto_periodo(idx_alvo):
        top_horas_brutas = pd.to_numeric(horas.loc[idx_alvo], errors='coerce').astype(int).tolist()
        top_pld_brutas = pd.to_numeric(pld.loc[idx_alvo], errors='coerce').astype(float).tolist()
        top_list = [(i, h, p) for i, (h, p) in enumerate(zip(top_horas_brutas, top_pld_brutas), 1)]
        top_sorted = sorted(top_list, key=lambda x: x[1])

        texto = ""
        for rank, h, p in top_sorted:
            texto += f"    Hora {h} - R$ {p:.2f}\n"
        return texto, top_sorted

    texto_top3, top3_sorted = montar_texto_periodo(top3_idx)
    texto_top4, top4_sorted = montar_texto_periodo(top4_idx)

    # Simular 3 Horas
    consumo_3h = pd.Series([consumo_medio_dia] * len(pld), index=pld.index)
    consumo_3h.loc[top3_idx] = CONSUMO_REDUZIDO_MWM
    balanco_3h = geracao_mod - consumo_3h - cessao
    res_fin_3h = float((balanco_3h * pld).sum())
    soma_E_3h = float(consumo_3h.sum())
    denominador_3h = soma_E_3h + soma_F
    custo_mwm_3h = (res_fin_3h / denominador_flat) if denominador_flat else 0.0

    # Simular 4 Horas
    consumo_4h = pd.Series([consumo_medio_dia] * len(pld), index=pld.index)
    consumo_4h.loc[top4_idx] = CONSUMO_REDUZIDO_MWM
    balanco_4h = geracao_mod - consumo_4h - cessao
    res_fin_4h = float((balanco_4h * pld).sum())
    soma_E_4h = float(consumo_4h.sum())
    denominador_4h = soma_E_4h + soma_F
    custo_mwm_4h = (res_fin_4h / denominador_flat) if denominador_flat else 0.0

    ganho_3h = res_fin_3h - res_fin
    ganho_4h = res_fin_4h - res_fin
    ganho_extra_4a_hora = ganho_4h - ganho_3h

    # =============================
    # ATUALIZAR PLANILHA FÍSICA E RECALCULAR (VBS)
    # =============================
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        
        print("\nAtualizando a planilha física com o cenário de 4 horas...")
        wb = load_workbook(caminho)
        ws = wb[nome_aba]
        
        fonte_vermelha = Font(color="FF0000", bold=True)
        
        for idx in pld.index: 
            # O índice 5 no Pandas é a linha 6 no Excel
            linha_excel = idx + 1
            
            # Atualiza o consumo na coluna E (coluna 5)
            ws.cell(row=linha_excel, column=5).value = float(consumo_4h.loc[idx])
            
            # Se for uma das 4 horas mais caras, pinta a coluna B (Hora) e E (Consumo c/ redução) de vermelho
            if idx in top4_idx:
                ws.cell(row=linha_excel, column=2).font = fonte_vermelha
                ws.cell(row=linha_excel, column=5).font = fonte_vermelha
                
        wb.save(caminho)
        wb.close()
        print("   [v] Planilha atualizada e salva pelo Python!")
        
        # Chama o VBS para recalcular as fórmulas
        print("   [v] Rodando VBScript para recalcular fórmulas do arquivo Excel...")
        subprocess.run(["cscript.exe", "//nologo", VBS_SCRIPT, caminho], check=True)
        print("   [v] Planilha recalculada e consolidada com sucesso!")
        
    except Exception as e:
        print(f"[AVISO] Não foi possível formatar/recalcular a planilha física: {e}")

    # =============================
    # LÓGICA DE DATA-ALVO
    # =============================
    agora = datetime.now()
    if agora.hour >= 17:
        data_alvo = agora + timedelta(days=1)
    else:
        data_alvo = agora
    
    data_relatorio = data_alvo.strftime("%d/%m/%Y")

    # =============================
    # MENSAGEM FINAL (OUTPUT)
    # =============================
    icone_base = "🟢" if res_fin >= 0 else "🔴"
    icone_3h = "🟢" if res_fin_3h >= 0 else "🔴"
    icone_4h = "🟢" if res_fin_4h >= 0 else "🔴"

    msg_html = (
        "========================================\n"
        "<b>📊 RELATÓRIO DE MODULAÇÃO</b>\n"
        f"📅 {data_relatorio}\n\n"

        "<b>🔌 Cenário Operacional Simulado</b>\n"
        f"Consumo Médio: {consumo_medio_dia:.2f} MWm\n"
        f"Total Recurso: {to_num(df.iloc[1, col_B]):.2f} MWm\n\n"

        "<b>⚡ Análise do PLD</b>\n"
        f"📈 Máx: R$ {val_max:.2f} - Hora {hora_max}\n"
        f"📉 Mín: R$ {val_min:.2f} - Hora {hora_min}\n"
        f"⚖️ Méd: R$ {media_pld:.2f}\n\n"

        "⏱️ <b>Período de possível redução (Redução 3h):</b>\n"
        f"{texto_top3}\n"
        "⏱️ <b>Período de possível redução (Redução 4h):</b>\n"
        f"{texto_top4}\n"

        "<b>🧮 Ganhos vs Cenário Flat</b>\n"
        f"Reduzir 3 Horas (Ganho): + R$ {ganho_3h:,.2f}\n"
        f"Reduzir 4 Horas (Ganho): + R$ {ganho_4h:,.2f} (+ R$ {ganho_extra_4a_hora:,.2f} extra)\n\n"

        "<b>💰 Resultado Financeiro</b>\n"
        f"{icone_base} Financeiro Consumo Flat: R$ {res_fin:,.2f}"
        f" [R$ {custo_mwm_flat:.2f}/MWh]\n"
        f"{icone_3h} Financeiro c/ Redução 3h: R$ {res_fin_3h:,.2f}"
        f" [R$ {custo_mwm_3h:.2f}/MWh]\n"
        f"{icone_4h} Financeiro c/ Redução 4h: R$ {res_fin_4h:,.2f}"
        f" [R$ {custo_mwm_4h:.2f}/MWh]\n\n"

        "========================\n"
    )

    print("\n" + "=" * 40)
    texto_terminal = (msg_html
                      .replace("<b>", "").replace("</b>", "")
                      .replace("<i>", "").replace("</i>", "")
                      .replace("<code>", "").replace("</code>", ""))
    print(texto_terminal)
    print("=" * 40 + "\n")

    # =============================
    # GRÁFICO E ENVIO TELEGRAM
    # =============================
    destaques_sorted = sorted([(hora_min, val_min), (hora_max, val_max)], key=lambda x: x[0])
    gerar_grafico_pld(horas, pld, media_pld, TEMP_IMG, destaques_sorted, data_relatorio)
    
    if os.path.exists(TEMP_IMG):
        if enviar_telegram_com_foto(msg_html, TEMP_IMG, TOKEN_TELEGRAM, CHAT_ID):
            print("✅ Relatório e Gráfico enviados com sucesso!")
        else:
            print("❌ Falha no envio da Foto para o Telegram.")
            
    # Enviar o próprio arquivo Excel no Telegram
    if os.path.exists(caminho):
        if enviar_telegram_documento(caminho, TOKEN_TELEGRAM, CHAT_ID):
            print("✅ Planilha anexada e enviada com sucesso!")
        else:
            print("❌ Falha ao anexar a Planilha no Telegram.")

if __name__ == "__main__":
    main()