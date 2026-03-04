import os
import sys
import subprocess
import configparser
import time
from datetime import datetime, timedelta

# ==========================================
# UTILITÁRIOS E VALIDAÇÕES
# ==========================================
def imprimir_cabecalho(mensagem):
    print("\n" + "=" * 60)
    print(mensagem)
    print("=" * 60)

def arquivo_foi_modificado_hoje(pasta, extensao=".xlsx"):
    """Verifica se existe algum arquivo na pasta que foi criado/modificado hoje."""
    if not os.path.exists(pasta):
        return False

    hoje = datetime.now().date()
    for arquivo in os.listdir(pasta):
        if arquivo.lower().endswith(extensao):
            caminho = os.path.join(pasta, arquivo)
            data_modificacao = datetime.fromtimestamp(os.path.getmtime(caminho)).date()
            if data_modificacao == hoje:
                return True
    return False

def validar_pld_baixado(pasta_pld, target_date):
    """
    Abre os XLSX recém-baixados na pasta e procura ativamente a data alvo na coluna 'DIA'.
    Retorna True se achar.
    """
    try:
        import openpyxl
    except ImportError:
        print(" [ERRO] A biblioteca 'openpyxl' não está instalado.")
        return False

    for arquivo in os.listdir(pasta_pld):
        if arquivo.lower().endswith(".xlsx") and not arquivo.startswith("~$"):
            caminho = os.path.join(pasta_pld, arquivo)
            try:
                wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
                ws = wb.worksheets[0]

                header = [str(cell.value).strip().upper() for cell in ws[1] if cell.value]
                if "DIA" not in header:
                    continue

                idx_dia = header.index("DIA")
                encontrou = False

                for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
                    if row[idx_dia] is not None:
                        try:
                            if int(row[idx_dia]) == target_date.day:
                                encontrou = True
                                break
                        except:
                            pass

                wb.close()
                if encontrou:
                    return True

            except Exception as e:
                print(f" [AVISO] Não foi possível validar {arquivo}: {e}")
                
    return False

def executar_download_pld_com_validacao(script_path, max_tentativas, espera_segundos, pasta_pld):
    """Executa o script de PLD e em seguida valida o conteúdo."""
    agora = datetime.now()
    if agora.hour >= 17:
        target_date = (agora + timedelta(days=1)).date()
    else:
        target_date = agora.date()

    for tentativa in range(1, max_tentativas + 1):
        print(f" [Tentativa {tentativa}/{max_tentativas}] Baixando PLD...")
        try:
            subprocess.run([sys.executable, script_path], check=True)
            print(" [v] Script concluiu com sucesso. Validando dados baixados...")

            if validar_pld_baixado(pasta_pld, target_date):
                print(f" [+] SUCESSO! Dados do dia {target_date.strftime('%d/%m/%Y')} confirmados no XLSX.")
                return True
            else:
                print(f" [!] XLSX baixado, porém os dados do dia {target_date.strftime('%d/%m/%Y')} NÃO ESTÃO LÁ AINDA.")
                
        except subprocess.CalledProcessError as e:
            print(f" [AVISO] Falha ao tentar rodar o script de PLD (Erro {e.returncode}).")

        if tentativa < max_tentativas:
            print(f" [!] A CCEE ainda não disponibilizou os dados finais ou o site caiu.")
            print(f" [!] Aguardando {espera_segundos / 60:.1f} minutos para nova tentativa...")
            time.sleep(espera_segundos)
        else:
            print(" [ERRO FATAL] Todas as tentativas esgotadas. PLD indisponível hoje.")
            return False

def calcular_segundos_ate_horario(horario_str):
    """Calcula quantos segundos faltam de 'agora' até o 'horário' especificado."""
    agora = datetime.now()
    try:
        hora, minuto = map(int, horario_str.split(':'))
    except ValueError:
        print(f" [ERRO] Formato de horário inválido: {horario_str}. Usando 17:15 como padrão.")
        hora, minuto = 17, 15
        
    alvo = agora.replace(hour=hora, minute=minuto, second=0, microsecond=0)
    
    # Se o horário de hoje já passou, o alvo será amanhã neste mesmo horário
    if agora >= alvo:
        alvo += timedelta(days=1)
        
    return (alvo - agora).total_seconds()

# ==========================================
# ROTINA DIÁRIA
# ==========================================
def rotina_diaria_de_modulacao():
    """Esta é a função que será chamada todos os dias no horário agendado."""
    imprimir_cabecalho("INICIANDO A ROTINA DIÁRIA DE MODULAÇÃO")
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Lê o config a cada execução
    config_path = os.path.join(base_dir, "Configuracoes", "config.ini")
    config = configparser.ConfigParser()
    config.optionxform = str
    config.read(config_path, encoding='utf-8')

    max_tentativas_ccee = config.getint("ORQUESTRADOR", "MAX_TENTATIVAS_CCEE", fallback=20)
    tempo_espera_minutos = config.getint("ORQUESTRADOR", "TEMPO_ESPERA_MINUTOS", fallback=20)

    # Verifica Regras de Negócio antes de iniciar
    print("Verificando variáveis de negócio...")
    chaves = [
        ("TELEGRAM", "BOT_TOKEN"),
        ("TELEGRAM", "CHAT_ID"),
        ("REGRAS_NEGOCIO", "CONSUMO_MEDIO_MWM_MES"),
        ("REGRAS_NEGOCIO", "TOTAL_RECURSO_MES"),
        ("REGRAS_NEGOCIO", "CONSUMO_REDUZIDO_MWM"),
    ]

    vazios = [
        f"[{secao}] -> {chave}"
        for secao, chave in chaves
        if (not config.has_option(secao, chave)) or (not config.get(secao, chave).strip())
    ]

    if vazios:
        print(" [ERRO CRÍTICO] Preencha no 'config.ini':", ", ".join(vazios))
        print(" [BLOQUEIO] A rotina de hoje foi cancelada por falta de variáveis. Tentarei amanhã.")
        return # O "return" cancela a rotina de hoje, mas mantém o serviço rodando

    print("\n" + "=" * 60 + "\nINICIANDO CASCATA\n" + "=" * 60)

    # PASSO 1: Baixar PLD
    print(f"\n PASSO 1: Baixar PLD (Até {max_tentativas_ccee} tentativas)")
    script_pld = os.path.join(base_dir, "Src", "baixar_pld_ccee_sudeste_xlsx.py")
    pasta_pld = os.path.join(base_dir, "PLD_Horario_Sudeste")

    if not executar_download_pld_com_validacao(script_pld, max_tentativas_ccee, tempo_espera_minutos * 60, pasta_pld):
        print("\n[BLOQUEIO] Falha crítica ao obter o PLD do dia. A rotina de hoje foi abortada.")
        return

    # PASSO 2: Preencher Template
    print("\n PASSO 2: Preencher Template")
    script_gerar = os.path.join(base_dir, "Src", "gerar_modulacao_parada_diaria_v3.py")
    try:
        subprocess.run([sys.executable, script_gerar], check=True)
    except subprocess.CalledProcessError:
        print("\n[BLOQUEIO] Falha ao gerar a planilha. A rotina de hoje foi abortada.")
        return

    if not arquivo_foi_modificado_hoje(os.path.join(base_dir, "Planilha_Modulacao")):
        print("\n[BLOQUEIO] Planilha não encontrada na pasta de saída. Abortando hoje.")
        return

    # PASSO 3: Analisar Cenários e Enviar
    print("\n PASSO 3: Analisar Cenários e Enviar Telegram")
    script_notifica = os.path.join(base_dir, "Src", "NotificaCustoModulacao.py")
    try:
        subprocess.run([sys.executable, script_notifica], check=True)
    except subprocess.CalledProcessError:
        print("\n[ERRO] O relatório financeiro falhou ao ser enviado. Abortando hoje.")
        return

    imprimir_cabecalho("Processo concluído com sucesso!")


# ==========================================
# INICIALIZAÇÃO E LOOP DO SERVIÇO NATIVO
# ==========================================
def main():
    imprimir_cabecalho("SERVIÇO DE MODULAÇÃO INICIADO")
    base_dir = os.path.dirname(os.path.abspath(__file__))

    config_path = os.path.join(base_dir, "Configuracoes", "config.ini")
    if not os.path.exists(config_path):
        print(f"[FATAL] Arquivo 'config.ini' não encontrado em: {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.optionxform = str
    config.read(config_path, encoding='utf-8')

    # Configuração inicial obrigatória
    if "ORQUESTRADOR" not in config: config.add_section("ORQUESTRADOR")
    if not config.has_option("ORQUESTRADOR", "MAX_TENTATIVAS_CCEE"): config.set("ORQUESTRADOR", "MAX_TENTATIVAS_CCEE", "20")
    if not config.has_option("ORQUESTRADOR", "TEMPO_ESPERA_MINUTOS"): config.set("ORQUESTRADOR", "TEMPO_ESPERA_MINUTOS", "20")
    if not config.has_option("ORQUESTRADOR", "HORARIO_EXECUCAO"): config.set("ORQUESTRADOR", "HORARIO_EXECUCAO", "17:15")

    # Criação de Pastas
    pastas_necessarias = ["PLD_Horario_Sudeste", "Planilha_Modulacao", "Logs", "GraficosPLD"]
    for pasta in pastas_necessarias:
        os.makedirs(os.path.join(base_dir, pasta), exist_ok=True)

    template_path = os.path.join(base_dir, 'Templates', 'AAAA.MM.DD_Modulacao_Consumo e Cessao.xlsx')
    if not os.path.exists(template_path):
        print("\n[FATAL] Planilha Template não encontrada na pasta 'Templates'!")
        sys.exit(1)

    # Reescreve rotas
    if "DIRETORIOS" not in config: config.add_section("DIRETORIOS")
    config.set("DIRETORIOS", "PLD_HORARIO", os.path.join(base_dir, "PLD_Horario_Sudeste"))
    config.set("DIRETORIOS", "TEMPLATE_PLANILHA", template_path)
    config.set("DIRETORIOS", "SAIDA_PLANILHAS", os.path.join(base_dir, "Planilha_Modulacao"))
    config.set("DIRETORIOS", "DIRETORIO_LOGS", os.path.join(base_dir, "Logs"))
    config.set("DIRETORIOS", "DIRETORIO_NOTIFICACAO", base_dir)
    config.set("DIRETORIOS", "VBS_SCRIPT", os.path.join(base_dir, "Src", "recalcular_salvar_fechar.vbs"))
    config.set("DIRETORIOS", "TEMP_IMG", os.path.join(base_dir, "GraficosPLD", "grafico_pld.png"))

    with open(config_path, "w", encoding="utf-8") as f:
        config.write(f)

    print(f"\n[v] Ambiente validado e configurado.")

    # ===============================================
    # LOOP INFINITO
    # ===============================================
    while True:
        # Lê o horário do config.ini a cada volta
        # Permite que o usuário altere o horário de execução sem precisar reiniciar o serviço
        config.read(config_path, encoding='utf-8')
        horario = config.get("ORQUESTRADOR", "HORARIO_EXECUCAO", fallback="17:15")
        
        # Calcula quantos segundos faltam para chegar nesse horário
        segundos_espera = calcular_segundos_ate_horario(horario)
        horas_espera = segundos_espera / 3600
        
        print(f"\n[zZz] Aguardando execução. Hora Definida: {horas_espera:.2f} horas, exatamente às {horario}.")
        
        # O programa pausa aqui até chegar no horário definido, economizando recursos do sistema
        time.sleep(segundos_espera)
        
        # Quando acorda, executa a rotina
        rotina_diaria_de_modulacao()

if __name__ == "__main__":
    main()
